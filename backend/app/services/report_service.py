import csv
import io
from datetime import date, timedelta

from sqlalchemy import select, func
from sqlalchemy.orm import Session

from app.models.entry import LogEntry
from app.models.project import Project


def _week_bounds(anchor: date) -> tuple[date, date]:
    """Return (monday, sunday) for the ISO week containing `anchor`."""
    monday = anchor - timedelta(days=anchor.weekday())
    sunday = monday + timedelta(days=6)
    return monday, sunday


def _month_bounds(anchor: date) -> tuple[date, date]:
    """Return (first_day, last_day) for the calendar month containing `anchor`."""
    first = anchor.replace(day=1)
    # Next month first day minus one day
    if anchor.month == 12:
        last = date(anchor.year + 1, 1, 1) - timedelta(days=1)
    else:
        last = date(anchor.year, anchor.month + 1, 1) - timedelta(days=1)
    return first, last


def get_summary(db: Session, anchor: date) -> dict:
    anchor_str = anchor.isoformat()
    monday, sunday = _week_bounds(anchor)
    month_start, month_end = _month_bounds(anchor)

    monday_str = monday.isoformat()
    sunday_str = sunday.isoformat()
    month_start_str = month_start.isoformat()
    month_end_str = month_end.isoformat()

    # today_minutes
    today_result = db.execute(
        select(func.coalesce(func.sum(LogEntry.duration_minutes), 0)).where(
            LogEntry.entry_date == anchor_str
        )
    ).scalar_one()

    # week_minutes
    week_result = db.execute(
        select(func.coalesce(func.sum(LogEntry.duration_minutes), 0)).where(
            LogEntry.entry_date >= monday_str,
            LogEntry.entry_date <= sunday_str,
        )
    ).scalar_one()

    # month_minutes
    month_result = db.execute(
        select(func.coalesce(func.sum(LogEntry.duration_minutes), 0)).where(
            LogEntry.entry_date >= month_start_str,
            LogEntry.entry_date <= month_end_str,
        )
    ).scalar_one()

    # by_project — week totals grouped by project
    by_project_rows = db.execute(
        select(
            Project.id,
            Project.name,
            Project.color,
            func.coalesce(func.sum(LogEntry.duration_minutes), 0).label("week_minutes"),
        )
        .join(LogEntry, LogEntry.project_id == Project.id)
        .where(
            LogEntry.entry_date >= monday_str,
            LogEntry.entry_date <= sunday_str,
        )
        .group_by(Project.id, Project.name, Project.color)
        .order_by(Project.name)
    ).all()

    by_project = [
        {
            "project_id": row.id,
            "name": row.name,
            "color": row.color,
            "week_minutes": row.week_minutes,
        }
        for row in by_project_rows
    ]

    # daily_totals — one entry per day of the week (Mon to Sun), 0 if no entries
    daily_rows = db.execute(
        select(
            LogEntry.entry_date,
            func.sum(LogEntry.duration_minutes).label("minutes"),
        )
        .where(
            LogEntry.entry_date >= monday_str,
            LogEntry.entry_date <= sunday_str,
        )
        .group_by(LogEntry.entry_date)
    ).all()

    daily_map = {row.entry_date: row.minutes for row in daily_rows}

    daily_totals = []
    for i in range(7):
        day = monday + timedelta(days=i)
        day_str = day.isoformat()
        daily_totals.append({"date": day_str, "minutes": daily_map.get(day_str, 0)})

    return {
        "anchor_date": anchor_str,
        "today_minutes": today_result,
        "week_minutes": week_result,
        "month_minutes": month_result,
        "by_project": by_project,
        "daily_totals": daily_totals,
    }


def export_csv(
    db: Session,
    from_date: str | None = None,
    to_date: str | None = None,
    project_id: int | None = None,
) -> str:
    stmt = (
        select(
            LogEntry.entry_date,
            Project.name.label("project_name"),
            LogEntry.duration_minutes,
            LogEntry.start_time,
            LogEntry.end_time,
            LogEntry.notes,
        )
        .join(Project, Project.id == LogEntry.project_id)
    )

    if from_date is not None:
        stmt = stmt.where(LogEntry.entry_date >= from_date)
    if to_date is not None:
        stmt = stmt.where(LogEntry.entry_date <= to_date)
    if project_id is not None:
        stmt = stmt.where(LogEntry.project_id == project_id)

    stmt = stmt.order_by(LogEntry.entry_date, Project.name)

    rows = db.execute(stmt).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["date", "project", "duration_hours", "start_time", "end_time", "notes"])

    for row in rows:
        duration_hours = round(row.duration_minutes / 60, 4)
        writer.writerow([
            row.entry_date,
            row.project_name,
            duration_hours,
            row.start_time or "",
            row.end_time or "",
            row.notes or "",
        ])

    return output.getvalue()


def _round_to_quarter(hours: float) -> float:
    """Round to nearest 0.25 per timesheet convention."""
    return round(hours * 4) / 4


def generate_timesheet_xlsx(db: Session, month: date) -> bytes:
    """
    Fill in the timesheet template with hours from the database for a given month.
    Returns raw bytes of the xlsx file.
    """
    import io
    from pathlib import Path
    from openpyxl import load_workbook
    from sqlalchemy import func

    template_path = Path(__file__).parent.parent.parent / "storage" / "timesheet_template.xlsx"
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found at {template_path}")

    # Load employee settings
    settings_path = Path(__file__).parent.parent.parent / "storage" / "settings.json"
    import json
    settings = json.loads(settings_path.read_text(encoding="utf-8")) if settings_path.exists() else {}

    # Compute sheet start date: Monday of the week containing the 1st of the month
    first_of_month = month.replace(day=1)
    sheet_start = first_of_month - timedelta(days=first_of_month.weekday())  # Monday
    sheet_end = sheet_start + timedelta(days=36)  # 37 columns (C to AM)

    # Query: sum duration per project per day within the sheet range
    rows = db.execute(
        select(
            LogEntry.entry_date,
            LogEntry.project_id,
            Project.name,
            Project.code,
            func.sum(LogEntry.duration_minutes).label("total_minutes"),
        )
        .join(Project, Project.id == LogEntry.project_id)
        .where(
            LogEntry.entry_date >= sheet_start.isoformat(),
            LogEntry.entry_date <= sheet_end.isoformat(),
        )
        .group_by(LogEntry.entry_date, LogEntry.project_id, Project.name, Project.code)
        .order_by(Project.name, LogEntry.entry_date)
    ).all()

    # Build {project_id: {date_str: hours}}
    projects_data: dict[int, dict[str, float]] = {}
    project_info: dict[int, tuple[str, str]] = {}
    for row in rows:
        pid = row.project_id
        if pid not in projects_data:
            projects_data[pid] = {}
            project_info[pid] = (row.name, row.code or "")
        projects_data[pid][row.entry_date] = _round_to_quarter(row.total_minutes / 60)

    # Load and fill the template
    wb = load_workbook(template_path)
    ws = wb["Timesheet"]

    # Fill metadata
    ws["B3"] = settings.get("employee_name", "")
    ws["B4"] = settings.get("department", "")
    ws["B6"] = first_of_month  # Excel will recalculate date formulas from this

    # Clear project rows 12–31 (columns A, B, C–AM only; preserve AN SUM formulas)
    for excel_row in range(12, 32):
        ws.cell(row=excel_row, column=1).value = None  # A: project name
        ws.cell(row=excel_row, column=2).value = None  # B: code
        for col in range(3, 40):  # C (3) through AM (39)
            ws.cell(row=excel_row, column=col).value = None

    # Fill in project rows
    for i, (pid, day_data) in enumerate(projects_data.items()):
        if i >= 20:
            break
        excel_row = 12 + i
        name, code = project_info[pid]
        ws.cell(row=excel_row, column=1).value = name
        ws.cell(row=excel_row, column=2).value = code

        for date_str, hours in day_data.items():
            try:
                entry_date = date.fromisoformat(date_str)
            except ValueError:
                continue
            col_offset = (entry_date - sheet_start).days
            if 0 <= col_offset <= 36:
                ws.cell(row=excel_row, column=3 + col_offset).value = hours

    # Update supervisor name in the last row if present
    supervisor = settings.get("supervisor_name", "")
    if supervisor:
        # Find the supervisor row (search for "Supervisor Name:" in column A)
        for row in ws.iter_rows(min_col=1, max_col=1):
            cell = row[0]
            if cell.value and isinstance(cell.value, str) and "Supervisor Name:" in cell.value:
                cell.value = f"Supervisor Name: {supervisor}"
                break

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
